"""
Filedge GST — Selenium Automation
==================================
HOW TO RUN:
  1. Place this script in a folder with your Excel file
  2. Run:  python gst_checker.py
  3. A browser UI will open — upload Excel, pick Financial Year, click Start
  4. For each client, enter the CAPTCHA shown on the Search Taxpayer page
     and click SEARCH manually in the Chrome window
  5. Click "Next Client" anytime to skip to the next GSTIN
  6. Click "Stop Session" to pause — progress is saved
  7. Click "Resume Session" to continue from where you left off
  8. Download the report anytime — even mid-run

REQUIRED LIBRARIES:
  pip install selenium openpyxl pandas flask webdriver-manager pyinstaller
"""

import os, json, time, threading, traceback, sys
import pandas as pd
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template_string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import webbrowser
import tempfile

# Force UTF-8 output safely
try:
    if sys.stdout and sys.stdout.encoding:
        if sys.stdout.encoding.lower() != 'utf-8':
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except:
    pass

app = Flask(__name__)

# Works correctly both as .py and as frozen .exe
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_FILE  = os.path.join(BASE_DIR, "gst_report.xlsx")
SESSION_FILE = os.path.join(BASE_DIR, "gst_session.json")
CAPTCHA_WAIT = 120  # max seconds to wait for manual CAPTCHA entry per client

GSTIN_SEARCH_URL = "https://services.gst.gov.in/services/searchtp"

# Month order for GST financial year (April → March)
MONTHS = ["April","May","June","July","August","September",
          "October","November","December","January","February","March"]
# Short names for Excel columns
MONTHS_SHORT = ["APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC","JAN","FEB","MAR"]

# ── Runtime state ──
progress_log        = []
results_data        = []
is_done             = False
skip_flag           = False
stop_flag           = False
uploaded_df         = None
current_client      = {"name": "", "gstin": "", "index": 0}
progress            = {"done": 0, "total": 0}
session_active      = False
session_ever_started = False

# ─────────────────────────────────────────────
# HTML UI
# ─────────────────────────────────────────────
HTML_PAGE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Filedge GST</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', sans-serif; background: #f0f4f8; color: #333; }
  header { background: #1a3c6e; color: white; padding: 18px 32px; }
  header h1 { font-size: 20px; }
  header p  { font-size: 13px; opacity: 0.8; margin-top: 2px; }
  .container { max-width: 960px; margin: 36px auto; padding: 0 16px; }
  .card { background: white; border-radius: 10px; padding: 28px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); margin-bottom: 24px; }
  .card h2 { font-size: 16px; color: #1a3c6e; margin-bottom: 18px; border-bottom: 2px solid #e2e8f0; padding-bottom: 10px; }
  label { font-size: 13px; font-weight: 600; display: block; margin-bottom: 6px; color: #555; }
  input[type=file], select { width: 100%; padding: 10px 14px; border: 1.5px solid #d1d5db; border-radius: 7px; font-size: 14px; background: #fafafa; }
  .row { display: flex; gap: 16px; margin-bottom: 18px; }
  .row > div { flex: 1; }
  .btn { background: #1a3c6e; color: white; border: none; padding: 12px 28px; border-radius: 7px; font-size: 15px; cursor: pointer; width: 100%; font-weight: 600; transition: background 0.2s; }
  .btn:hover { background: #15305a; }
  .btn:disabled { background: #94a3b8; cursor: not-allowed; }
  .btn-green { background: #16a34a; } .btn-green:hover { background: #15803d; }
  .btn-red   { background: #dc2626; } .btn-red:hover   { background: #b91c1c; }
  .btn-sm { padding: 9px 18px; font-size: 13px; width: auto; border-radius: 6px; border: none; color: white; cursor: pointer; font-weight: 600; }

  #control-bar { display:none; align-items:center; gap:10px; background:#1e293b; border-radius:9px; padding:14px 18px; margin-bottom:14px; flex-wrap:wrap; }
  #control-bar .client-info { color:#94a3b8; font-size:13px; flex:1; min-width:160px; }
  #control-bar .client-info b { color:#f1f5f9; font-size:14px; }
  .ctrl-btn-group { display:flex; gap:8px; flex-wrap:wrap; }

  #resume-banner { display:none; background:#fef9c3; border:1.5px solid #fbbf24; border-radius:9px; padding:14px 18px; margin-bottom:20px; }
  #resume-banner p { font-size:13.5px; color:#92400e; margin-bottom:10px; }
  #resume-banner .btn-row { display:flex; gap:10px; }

  #log-box { background:#0f172a; color:#86efac; font-family:monospace; font-size:12px; padding:16px; border-radius:8px; height:320px; overflow-y:auto; display:none; white-space:pre-wrap; line-height:1.7; }
  .badge { display:inline-block; padding:2px 10px; border-radius:99px; font-size:11px; font-weight:600; }
  .filed   { background:#dcfce7; color:#15803d; }
  .pending { background:#fef9c3; color:#92400e; }
  .skipped { background:#e0f2fe; color:#0369a1; }
  .stopped { background:#f3e8ff; color:#6b21a8; }
  .failed  { background:#fee2e2; color:#b91c1c; }
  table { width:100%; border-collapse:collapse; font-size:12px; }
  th { background:#1a3c6e; color:white; padding:8px 10px; text-align:center; white-space:nowrap; }
  th.left { text-align:left; }
  td { padding:7px 10px; border-bottom:1px solid #e2e8f0; text-align:center; white-space:nowrap; }
  td.left { text-align:left; }
  tr:hover td { background:#f8fafc; }
  #preview-section, #result-section { display:none; }
  .spinner { display:inline-block; width:13px; height:13px; border:2px solid rgba(255,255,255,0.4); border-top-color:white; border-radius:50%; animation:spin 0.7s linear infinite; vertical-align:middle; margin-right:6px; }
  @keyframes spin { to { transform:rotate(360deg); } }
  .note { font-size:12px; color:#64748b; margin-top:6px; }
  .col-fmt { background:#f1f5f9; border-radius:6px; padding:10px 14px; font-size:12.5px; margin-bottom:16px; }
  .col-fmt span { display:inline-block; background:#1a3c6e; color:white; padding:2px 8px; border-radius:4px; margin:2px; font-size:11px; }
  .progress-wrap { background:#e2e8f0; border-radius:99px; height:8px; margin:10px 0; overflow:hidden; }
  .progress-bar  { background:#16a34a; height:100%; border-radius:99px; transition:width 0.4s; width:0%; }
  #progress-section { display:none; }

  /* Floating download button — always visible once 1 client done */
  #dl-fab { display:none; position:fixed; bottom:28px; right:28px; z-index:9999;
    background:#16a34a; color:white; border:none; border-radius:50px; padding:14px 24px;
    font-size:14px; font-weight:700; cursor:pointer; box-shadow:0 4px 20px rgba(0,0,0,0.25);
    align-items:center; gap:10px; transition:background 0.2s, transform 0.15s; }
  #dl-fab:hover { background:#15803d; transform:translateY(-2px); }
  #dl-fab .badge-count { background:rgba(255,255,255,0.25); border-radius:99px; padding:2px 10px; font-size:12px; }
</style>
</head>
<body>
<header style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px;">
  <div>
    <h1>🧾 Filedge GST</h1>
    <p>Automated bulk monthly filing status retrieval via GST Portal</p>
  </div>
  <button onclick="quitApp()"
    style="background:#dc2626;color:white;border:none;padding:9px 20px;border-radius:7px;
           font-size:13px;font-weight:700;cursor:pointer;white-space:nowrap;">
    ✕ Quit App
  </button>
</header>

<div class="container">

  <div id="resume-banner">
    <p id="resume-info">📌 A previous session was stopped. You can resume from where you left off.</p>
    <div class="btn-row">
      <button class="btn-sm btn-green" onclick="resumeSession()">↩ Resume Previous Session</button>
      <button class="btn-sm btn-red"   onclick="clearSession()">🗑 Discard &amp; Start Fresh</button>
    </div>
  </div>

  <!-- STEP 1 -->
  <div class="card">
    <h2>Step 1 — Upload Client Excel &amp; Select Financial Year</h2>
    <div class="col-fmt">
      <b>Required Excel Columns:</b><br>
      <span>GSTIN</span> <span>Client Name</span>
    </div>
    <br>
    <button class="btn btn-green" onclick="downloadTemplate()">
    ⬇ Download Excel Template
    </button>
    <br><br>
    <div class="row">
      <div>
        <label>Upload Excel File (.xlsx)</label>
        <input type="file" id="excel-file" accept=".xlsx">
      </div>
      <div>
        <label>Financial Year</label>
        <select id="fin-year">
          <option value="2026-2027">2026-2027</option>
          <option value="2025-2026" selected>2025-2026</option>
          <option value="2024-2025">2024-2025</option>
          <option value="2023-2024">2023-2024</option>
          <option value="2022-2023">2022-2023</option>
          <option value="2021-2022">2021-2022</option>
        </select>
      </div>
    </div>
    <button class="btn" onclick="previewExcel()">Preview Excel</button>
  </div>

  <!-- STEP 2 -->
  <div class="card" id="preview-section">
    <h2>Step 2 — Preview Clients</h2>
    <div id="preview-table-wrap"></div>
    <br>
    <p class="note">✅ Verify the data above before starting.</p>
    <br>
    <button class="btn" id="start-btn" onclick="startAutomation()">▶ Start Automation</button>
  </div>

  <!-- STEP 3 -->
  <div class="card" id="progress-section">
    <h2>Step 3 — Live Progress</h2>
    <div id="control-bar">
      <div class="client-info">
        <div style="font-size:11px;color:#64748b;margin-bottom:2px">CURRENTLY PROCESSING</div>
        <b id="current-client-name">—</b>
        <span style="color:#64748b;margin-left:8px;font-size:12px" id="current-client-gstin"></span>
      </div>
      <div class="ctrl-btn-group">
        <button class="btn-sm" id="next-btn" style="background:#f59e0b;color:#1c1917" onclick="skipToNext()">⏭ Next Client</button>
        <button class="btn-sm" id="stop-btn" style="background:#7c3aed" onclick="stopSession()">⏸ Stop Session</button>
        <button class="btn-sm" id="dl-inline-btn" style="background:#16a34a;display:none" onclick="downloadReport()">⬇ Download (<span id="dl-inline-count">0</span>)</button>
      </div>
    </div>
    <div style="display:flex;justify-content:space-between;font-size:12px;color:#64748b;margin-bottom:4px">
      <span id="progress-label">0 / 0 clients</span>
      <span id="progress-pct">0%</span>
    </div>
    <div class="progress-wrap"><div class="progress-bar" id="progress-bar"></div></div>
    <br>
    <p class="note" style="margin-bottom:10px">
      ⚠️ Watch the Chrome window — for each client, enter the CAPTCHA on the Search Taxpayer page and click SEARCH.<br>
      Click <b>⏸ Stop Session</b> to pause and save. Click <b>⏭ Next Client</b> to skip.
    </p>
    <div id="log-box"></div>
  </div>

  <!-- STEP 4 -->
  <div class="card" id="result-section">
    <h2>Step 4 — Results</h2>
    <div id="result-table-wrap" style="overflow-x:auto;"></div>
    <br>
    <button class="btn btn-green" onclick="downloadReport()">⬇ Download Excel Report</button>
  </div>

</div>

<!-- FLOATING DOWNLOAD FAB -->
<button id="dl-fab" onclick="downloadReport()">
  ⬇ Download Report
  <span class="badge-count" id="dl-count">0 clients</span>
</button>

<script>
let logInterval = null;
let lastLogLen  = 0;

window.onload = async () => {
  const r = await fetch('/session_status');
  const d = await r.json();
  if (d.has_session) {
    document.getElementById('resume-banner').style.display = 'block';
    document.getElementById('resume-info').textContent =
      `📌 Previous session stopped at client ${d.done} of ${d.total} (${d.fin_year}). Resume from client ${d.next_index + 1}.`;
  }
  // Show FAB if a partial report from a previous run exists
  await checkDownloadReady();
};

async function checkDownloadReady() {
  try {
    const r = await fetch('/log');
    const d = await r.json();
    updateDownloadUI(d.download_ready, d.download_count);
  } catch(e) {}
}

function updateDownloadUI(ready, count) {
  if (ready && count > 0) {
    // Floating FAB
    const fab = document.getElementById('dl-fab');
    fab.style.display = 'flex';
    document.getElementById('dl-count').textContent = count + (count === 1 ? ' client' : ' clients');
    // Inline button inside control bar
    const inlineBtn = document.getElementById('dl-inline-btn');
    if (inlineBtn) {
      inlineBtn.style.display = 'inline-block';
      document.getElementById('dl-inline-count').textContent = count;
    }
  }
}

async function previewExcel() {
  const file = document.getElementById('excel-file').files[0];
  if (!file) { alert('Please select an Excel file first.'); return; }
  const btn = event.target;
  btn.disabled = true; btn.textContent = 'Loading...';
  try {
    const fd = new FormData();
    fd.append('file', file);
    const r = await fetch('/upload', { method:'POST', body:fd });
    const d = await r.json();
    if (d.error) { alert('Upload error: ' + d.error); return; }
    renderPreview(d.columns, d.rows);
    document.getElementById('preview-section').style.display = 'block';
    document.getElementById('preview-section').scrollIntoView({ behavior:'smooth' });
  } catch(e) {
    alert('Error: ' + e);
  } finally {
    btn.disabled = false; btn.textContent = 'Preview Excel';
  }
}

function renderPreview(cols, rows) {
  let html = '<div style="overflow-x:auto"><table><tr>';
  cols.forEach(c => html += `<th>${c}</th>`);
  html += '</tr>';
  rows.slice(0,10).forEach(row => {
    html += '<tr>';
    cols.forEach(c => html += `<td>${row[c] ?? ''}</td>`);
    html += '</tr>';
  });
  if (rows.length > 10)
    html += `<tr><td colspan="${cols.length}" style="text-align:center;color:#64748b">...and ${rows.length-10} more rows</td></tr>`;
  html += '</table></div>';
  document.getElementById('preview-table-wrap').innerHTML = html;
}

async function startAutomation(resumeFrom = 0) {
  const finYear = document.getElementById('fin-year').value;
  document.getElementById('start-btn').disabled = true;
  document.getElementById('start-btn').innerHTML = '<span class="spinner"></span>Running...';
  document.getElementById('progress-section').style.display = 'block';
  document.getElementById('log-box').style.display = 'block';
  document.getElementById('control-bar').style.display = 'flex';
  document.getElementById('resume-banner').style.display = 'none';
  document.getElementById('progress-section').scrollIntoView({ behavior:'smooth' });
  await fetch('/start', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({ fin_year:finYear, resume_from:resumeFrom })
  });
  lastLogLen = 0;
  logInterval = setInterval(pollLog, 1500);
}

async function resumeSession() {
  const r = await fetch('/session_status');
  const d = await r.json();
  if (!d.has_session) { alert('No saved session found.'); return; }
  const file = document.getElementById('excel-file').files[0];
  if (!file) { alert('Please re-upload the same Excel file to resume.'); return; }
  const fd = new FormData();
  fd.append('file', file);
  await fetch('/upload', { method:'POST', body:fd });
  document.getElementById('fin-year').value = d.fin_year;
  document.getElementById('progress-section').style.display = 'block';
  document.getElementById('log-box').style.display = 'block';
  document.getElementById('control-bar').style.display = 'flex';
  document.getElementById('resume-banner').style.display = 'none';
  document.getElementById('progress-section').scrollIntoView({ behavior:'smooth' });
  await fetch('/start', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({ fin_year:d.fin_year, resume_from:d.next_index })
  });
  lastLogLen = 0;
  logInterval = setInterval(pollLog, 1500);
}

async function clearSession() {
  await fetch('/clear_session', { method:'POST' });
  document.getElementById('resume-banner').style.display = 'none';
}

async function pollLog() {
  try {
    const r = await fetch('/log');
    const d = await r.json();
    const box = document.getElementById('log-box');
    if (d.log.length > lastLogLen) {
      d.log.slice(lastLogLen).forEach(line => { box.innerHTML += line + '\\n'; });
      lastLogLen = d.log.length;
      box.scrollTop = box.scrollHeight;
    }
    if (d.current_client) {
      document.getElementById('current-client-name').textContent  = d.current_client.name  || '—';
      document.getElementById('current-client-gstin').textContent = d.current_client.gstin || '';
    }
    if (d.progress) {
      const done = d.progress.done, total = d.progress.total;
      const pct = total ? Math.round((done/total)*100) : 0;
      document.getElementById('progress-label').textContent = `${done} / ${total} clients`;
      document.getElementById('progress-pct').textContent   = pct + '%';
      document.getElementById('progress-bar').style.width   = pct + '%';
    }
    // Show download buttons as soon as 1 client is done
    updateDownloadUI(d.download_ready, d.download_count);
    if (d.done || d.stopped) {
      clearInterval(logInterval);
      document.getElementById('next-btn').disabled = true;
      document.getElementById('stop-btn').disabled = true;
      if (d.stopped) {
        const s  = await fetch('/session_status');
        const sd = await s.json();
        if (sd.has_session) {
          document.getElementById('resume-info').textContent =
            `📌 Session paused at client ${sd.done} of ${sd.total}. Re-upload Excel and click Resume.`;
          document.getElementById('resume-banner').style.display = 'block';
          document.getElementById('resume-banner').scrollIntoView({ behavior:'smooth' });
        }
      }
      if (d.done) loadResults();
    }
  } catch(e) { console.error('pollLog error:', e); }
}

async function skipToNext() {
  document.getElementById('next-btn').disabled = true;
  document.getElementById('next-btn').textContent = '⏭ Skipping...';
  await fetch('/skip', { method:'POST' });
  setTimeout(() => {
    document.getElementById('next-btn').disabled = false;
    document.getElementById('next-btn').textContent = '⏭ Next Client';
  }, 3000);
}

async function stopSession() {
  if (!confirm('Stop and save session? You can resume later.')) return;
  await fetch('/stop', { method:'POST' });
  clearInterval(logInterval);
}

async function loadResults() {
  const r = await fetch('/results');
  const d = await r.json();
  renderResults(d.results);
  document.getElementById('result-section').style.display = 'block';
  document.getElementById('result-section').scrollIntoView({ behavior:'smooth' });
}

function renderResults(rows) {
  if (!rows.length) { document.getElementById('result-table-wrap').innerHTML = '<p>No results yet.</p>'; return; }
  const months = ['APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC','JAN','FEB','MAR'];
  let html = '<table><thead>';
  html += '<tr><th class="left" rowspan="2">GSTIN</th><th class="left" rowspan="2">Client Name</th>';
  html += '<th colspan="12">GSTR-3B</th><th colspan="12">GSTR-1/IFF</th></tr>';
  html += '<tr>';
  months.forEach(m => html += `<th>${m}</th>`);
  months.forEach(m => html += `<th>${m}</th>`);
  html += '</tr></thead><tbody>';
  rows.forEach(row => {
    html += '<tr>';
    html += `<td class="left">${row['GSTIN'] || ''}</td>`;
    html += `<td class="left">${row['Client Name'] || ''}</td>`;
    months.forEach(m => {
      const val = row['3B_' + m] || 'Pending';
      const cls = val === 'Filed' ? 'filed' : val === 'Pending' ? 'pending' : val === 'Skipped' ? 'skipped' : 'stopped';
      html += `<td><span class="badge ${cls}">${val}</span></td>`;
    });
    months.forEach(m => {
      const val = row['R1_' + m] || 'Pending';
      const cls = val === 'Filed' ? 'filed' : val === 'Pending' ? 'pending' : val === 'Skipped' ? 'skipped' : 'stopped';
      html += `<td><span class="badge ${cls}">${val}</span></td>`;
    });
    html += '</tr>';
  });
  html += '</tbody></table>';
  document.getElementById('result-table-wrap').innerHTML = html;
}

function downloadReport() { window.location.href = '/download'; }
function downloadTemplate() { window.location.href = '/download_template'; }
async function quitApp() {
  if (!confirm('Close GST Checker completely?')) return;
  try { await fetch('/quit', { method:'POST' }); } catch(e) {}
  document.body.innerHTML = '<div style="display:flex;flex-direction:column;align-items:center;' +
    'justify-content:center;height:100vh;font-family:Segoe UI,sans-serif;background:#f0f4f8;color:#1a3c6e;">' +
    '<div style="font-size:48px;margin-bottom:16px;">✅</div>' +
    '<h2>GST Checker has closed.</h2>' +
    '<p style="color:#64748b;margin-top:8px;">You can now close this tab and delete the .exe freely.</p></div>';
}
</script>
</body>
</html>
"""

# ─────────────────────────────────────────────
# SESSION HELPERS
# ─────────────────────────────────────────────
def save_session(next_index, fin_year, done, total, partial_results):
    with open(SESSION_FILE, "w") as f:
        json.dump({"next_index":next_index,"fin_year":fin_year,"done":done,
                   "total":total,"partial_results":partial_results,
                   "saved_at":datetime.now().strftime("%Y-%m-%d %H:%M:%S")}, f, indent=2)
    log(f"💾 Session saved. Next resume from client {next_index+1}.")

def load_session():
    if not os.path.exists(SESSION_FILE): return None
    with open(SESSION_FILE) as f: return json.load(f)

def clear_session_file():
    if os.path.exists(SESSION_FILE): os.remove(SESSION_FILE)

# ─────────────────────────────────────────────
# FLASK ROUTES
# ─────────────────────────────────────────────
@app.route("/")
def index(): return render_template_string(HTML_PAGE)

@app.route("/upload", methods=["POST"])
def upload():
    global uploaded_df
    f = request.files.get("file")
    if not f: return jsonify({"error":"No file uploaded"})
    try:
        import tempfile as _tf
        tmp = _tf.NamedTemporaryFile(delete=False, suffix=".xlsx")
        f.save(tmp.name)
        tmp.close()
        uploaded_df = pd.read_excel(tmp.name, engine="openpyxl")
        uploaded_df.columns = uploaded_df.columns.str.strip()
        rows = uploaded_df.fillna("").astype(str).to_dict(orient="records")
        os.unlink(tmp.name)
        return jsonify({"columns": list(uploaded_df.columns), "rows": rows})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc()})

@app.route("/start", methods=["POST"])
def start():
    global skip_flag,stop_flag,progress_log,results_data,is_done,session_active,session_ever_started
    data        = request.get_json()
    fin_year    = data.get("fin_year","2025-2026")
    resume_from = int(data.get("resume_from",0))
    skip_flag   = False
    stop_flag   = False
    is_done     = False
    session_active      = True
    session_ever_started = True
    if resume_from > 0:
        s = load_session()
        results_data = s.get("partial_results",[]) if s else []
        progress_log = []
        log(f"↩  Resuming from client {resume_from+1} — {len(results_data)} result(s) already saved.")
    else:
        progress_log = []
        results_data = []
    threading.Thread(target=run_automation, args=(fin_year,resume_from), daemon=True).start()
    return jsonify({"status":"started"})

@app.route("/skip", methods=["POST"])
def skip():
    global skip_flag
    skip_flag = True
    log("⏭  [USER] Skipping to next client...")
    return jsonify({"status":"ok"})

@app.route("/stop", methods=["POST"])
def stop():
    global stop_flag
    stop_flag = True
    log("⏸  [USER] Stop requested — saving session...")
    return jsonify({"status":"ok"})

@app.route("/session_status")
def session_status():
    s = load_session()
    if s: return jsonify({"has_session":True,"next_index":s["next_index"],
                          "fin_year":s["fin_year"],"done":s["done"],"total":s["total"]})
    return jsonify({"has_session":False})

@app.route("/clear_session", methods=["POST"])
def clear_session_route():
    clear_session_file()
    return jsonify({"status":"cleared"})

@app.route("/quit", methods=["POST"])
def quit_app():
    def _exit():
        time.sleep(0.6)
        os._exit(0)
    threading.Thread(target=_exit, daemon=True).start()
    return jsonify({"status":"bye"})

@app.route("/log")
def log_route():
    ready = os.path.exists(OUTPUT_FILE) and len(results_data) > 0
    return jsonify({
        "log"           : progress_log,
        "done"          : is_done,
        "stopped"       : stop_flag and session_ever_started and not is_done and not session_active,
        "current_client": current_client,
        "progress"      : progress,
        "download_ready": ready,
        "download_count": len(results_data),
    })

@app.route("/results")
def results_route(): return jsonify({"results":results_data})

@app.route("/download")
def download():
    try:
        import io
        if not os.path.exists(OUTPUT_FILE):
            return "Report not ready yet.", 404
        with open(OUTPUT_FILE, "rb") as f:
            buf = io.BytesIO(f.read())
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name="gst_report.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return f"Download error: {str(e)}", 500
@app.route("/download_template")
def download_template():
    try:
        import io, openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clients"
        headers = ["Client Name", "GSTIN"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor="1a3c6e")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = 26
        ws.row_dimensions[1].height = 20
        ws.append(["", ""])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name="gst_template.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback
        return "Error: " + str(e) + "\n" + traceback.format_exc(), 500

# ─────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────
def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    progress_log.append(f"[{ts}] {msg}")
    try:
        print(msg)
    except UnicodeEncodeError:
        # Strip non-ASCII and retry — keeps Windows charmap consoles happy
        print(msg.encode('ascii', errors='replace').decode('ascii'))

def should_skip(): return skip_flag or stop_flag

# ─────────────────────────────────────────────
# SELENIUM HELPERS
# ─────────────────────────────────────────────
def init_driver():
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_experimental_option("excludeSwitches",["enable-automation"])
    opts.add_experimental_option("useAutomationExtension",False)
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=opts
    )
    driver.execute_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
    return driver

def js_click(driver, element):
    """Click via JavaScript — bypasses Angular event interception."""
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
    time.sleep(0.3)
    driver.execute_script("arguments[0].click();", element)

def click_button_by_text(driver, label, timeout=15):
    """Find a button whose visible text exactly matches label and JS-click it."""
    label_up = label.strip().upper()
    end = time.time() + timeout
    while time.time() < end:
        btns = driver.find_elements(By.XPATH, "//button")
        for btn in btns:
            try:
                if btn.text.strip().upper() == label_up and btn.is_displayed():
                    js_click(driver, btn)
                    return True
            except:
                continue       
        time.sleep(0.5)
    # Fallback: contains match
    btns = driver.find_elements(By.XPATH, "//button")
    for btn in btns:
        if label_up in btn.text.strip().upper() and btn.is_displayed():
            js_click(driver, btn)
            return True
    raise Exception(f"Button '{label}' not found on page")

# ─────────────────────────────────────────────
# FETCH FILING STATUS — month-wise for GSTR-1 and GSTR-3B
# ─────────────────────────────────────────────
def fetch_filing_status(driver, gstin, fin_year):
    """
    Returns dict like:
      { "3B": {"April":"Filed","May":"Pending",...},
        "R1": {"April":"Filed","May":"Pending",...} }
    """
    empty = {m: "Pending" for m in MONTHS}
    if should_skip():
        label = "Skipped" if skip_flag else "Stopped"
        return {"3B":{m:label for m in MONTHS}, "R1":{m:label for m in MONTHS}}
    try:
        log("  🔍 Opening Search Taxpayer page...")
        driver.get(GSTIN_SEARCH_URL)

        # Wait for page to properly load
        inp = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((
                By.XPATH,
                "//input[@placeholder='Enter GSTIN/UIN of the Taxpayer' "
                "or contains(@name,'gstin') "
                "or contains(@id,'gstin')]"
            ))
        )
        time.sleep(1)
        if should_skip(): return _skipped_result()

        # ── Fill GSTIN ────────────────────────────────────────────────
        inp.clear(); inp.send_keys(gstin)
        time.sleep(0.5)
        if should_skip(): return _skipped_result()

        # ── Wait for user to enter CAPTCHA and click SEARCH ───────────
        log("  ✏️  GSTIN filled. Enter the CAPTCHA and click SEARCH in the browser...")
        log(f"  ⏳ Waiting up to {CAPTCHA_WAIT}s — click ⏭ Next or ⏸ Stop anytime.")
        elapsed = 0
        found = False
        while elapsed < CAPTCHA_WAIT:
            if stop_flag: log("  ⏸  Stopped during CAPTCHA."); return _skipped_result()
            if skip_flag: log("  ⏭  Skipped during CAPTCHA."); return _skipped_result()
            page = driver.page_source.lower()
            if "show filing table" in page or "filing details" in page or "no records found" in page:
                found = True
                log("  ✅ Taxpayer details loaded.")
                break
            time.sleep(1); elapsed += 1

        if not found:
            log("  ⏳ CAPTCHA / search timeout.")
            return {"3B": {m:"Captcha Timeout" for m in MONTHS}, "R1": {m:"Captcha Timeout" for m in MONTHS}}

        # ── Scroll to bottom — reveals SHOW FILING TABLE ──────────────
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        if should_skip(): return _skipped_result()

        # Log all buttons so we can debug if needed
        all_btns = driver.find_elements(By.XPATH, "//button")
        log(f"  🔎 Buttons on page: {[b.text.strip() for b in all_btns if b.text.strip()]}")

        # ── Click SHOW FILING TABLE ───────────────────────────────────
        log("  🖱️  Clicking SHOW FILING TABLE...")
        click_button_by_text(driver, "SHOW FILING TABLE")
        time.sleep(3)   # Angular renders the dropdown after this
        if should_skip(): return _skipped_result()

        # ── Select Financial Year ─────────────────────────────────────
        log(f"  📅 Selecting Financial Year: {fin_year}...")
        fy_selected = False
        try:
            fy_el = WebDriverWait(driver,12).until(
                EC.visibility_of_element_located((By.XPATH,"//select"))
            )
            sel = Select(fy_el)
            options = [o.text.strip() for o in sel.options]
            log(f"  📋 Dropdown options: {options}")
            if fin_year in options:
                sel.select_by_visible_text(fin_year)
                fy_selected = True
            else:
                for opt in options:
                    if opt.startswith(fin_year[:4]):
                        sel.select_by_visible_text(opt)
                        log(f"  📅 Matched: {opt}")
                        fy_selected = True; break
        except Exception as ey:
            log(f"  ⚠️  Dropdown error: {ey}")
        if not fy_selected:
            log("  ⚠️  Could not select year — using portal default.")
        time.sleep(1)
        if should_skip(): return _skipped_result()

        # ── Click SEARCH next to year dropdown ────────────────────────
        # The SEARCH button for filing table comes AFTER the <select>.
        # Use JS click directly — most reliable on Angular pages.
        log("  🖱️  Clicking SEARCH (filing table)...")
        search_clicked = False
        try:
            # Find all buttons, pick the one labelled SEARCH that is visible
            # and comes after the select element in DOM order
            all_btns = driver.find_elements(By.XPATH, "//select/following::button")
            for btn in all_btns:
                if btn.text.strip().upper() == "SEARCH" and btn.is_displayed() and btn.is_enabled():
                    js_click(driver, btn)
                    log("  ✅ Clicked SEARCH (filing) via JS.")
                    search_clicked = True
                    break
        except Exception as es:
            log(f"  ⚠️  SEARCH button error: {es}")

        if not search_clicked:
            # Hard fallback: click any visible SEARCH button
            all_btns = driver.find_elements(By.XPATH,"//button")
            for btn in reversed(all_btns):   # last one on page = the filing search
                if btn.text.strip().upper() == "SEARCH" and btn.is_displayed() and btn.is_enabled():
                    js_click(driver, btn)
                    log("  ✅ Clicked last SEARCH button (fallback).")
                    search_clicked = True; break

        if not search_clicked:
            log("  ⚠️  Could not click SEARCH — results may already be showing.")

        # Wait until filing table section appears
        log("  ⏳ Waiting for filing section to load...")

        table_loaded = False
        wait_start = time.time()

        while time.time() - wait_start < 25:

            if should_skip():
                return _skipped_result()

            page = driver.page_source.lower()

            # Filing section visible
            if (
                "gstr3b" in page or
                "gstr-1" in page or
                "filing details" in page or
                "no records found" in page
            ):
                table_loaded = True
                break

            time.sleep(1)

        if table_loaded:
            log("  ✅ Filing section loaded.")
        else:
            log("  ⚠️ Filing section load timeout — attempting scrape anyway.")

        # Small stabilization delay
        time.sleep(2)

        # ── Scrape tables ─────────────────────────────────────────────
        result = scrape_filing_tables(driver)
        log(f"  📊 GSTR-3B: {sum(1 for v in result['3B'].values() if v=='Filed')} months Filed | "
            f"GSTR-1: {sum(1 for v in result['R1'].values() if v=='Filed')} months Filed")
        return result

    except Exception as e:
        log(f"  ❌ fetch_filing_status error: {e}")
        log(traceback.format_exc())
        return {"3B": dict(empty), "R1": dict(empty)}

def _skipped_result():
    label = "Skipped" if skip_flag else "Stopped"
    return {"3B":{m:label for m in MONTHS}, "R1":{m:label for m in MONTHS}}

def scrape_filing_tables(driver):
    """
    Parse the two tables on the search result page:
      'Filing details for GSTR3B'   → 3B
      'Filing details for GSTR-1/IFF' → R1
    Each table has columns: Financial Year | Tax Period | Date of filing | Status
    Returns month→status dict for each return type.
    """
    result = {
        "3B": {m:"Pending" for m in MONTHS},
        "R1": {m:"Pending" for m in MONTHS},
    }

    page_src = driver.page_source.lower()
    if "no records found" in page_src:
        log("  ℹ️  'No Records Found' — all months Pending.")
        return result

    # Map: section heading keyword → result key
    sections = [
        ("GSTR3B",   "3B"),
        ("GSTR-1",   "R1"),
        ("GSTR1",    "R1"),   # alternative label
    ]

    for kw, key in sections:
        if result[key] != {m:"Pending" for m in MONTHS}:
            continue  # already filled this section

        # Find the heading element that contains the keyword
        headings = driver.find_elements(By.XPATH,
            f"//*[contains(text(),'{kw}')]")
        for heading in headings:
            try:
                # The table immediately follows the heading
                table = heading.find_element(By.XPATH, "following::table[1]")
                rows  = table.find_elements(By.TAG_NAME, "tr")
                found_any = False
                for row in rows:
                    tds = row.find_elements(By.TAG_NAME, "td")
                    if len(tds) < 4:
                        continue
                    # Columns: Financial Year | Tax Period | Date of filing | Status
                    tax_period = tds[1].text.strip()   # e.g. "March", "April"
                    status     = tds[3].text.strip()   # e.g. "Filed"
                    if tax_period in MONTHS:
                        result[key][tax_period] = status if status else "Pending"
                        found_any = True
                if found_any:
                    log(f"  ✅ Scraped {kw}: {result[key]}")
                    break
            except Exception as ep:
                log(f"  ⚠️  scrape error for {kw}: {ep}")
                continue

    return result

# ─────────────────────────────────────────────
# EXCEL SAVE — month-wise structure matching the output template
# ─────────────────────────────────────────────
def save_report_now():
    """
    Writes results_data to Excel with the structure:
    GSTIN | Client Name | GSTR-3B (12 months) | GSTR-1/IFF (12 months)
    """
    try:
        if not results_data:
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GST Filing Status"

        # ── Styles ──
        hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        hdr_fill_b = PatternFill("solid", fgColor="1a3c6e")   # dark blue — main header
        hdr_fill_g = PatternFill("solid", fgColor="1e6e3c")   # dark green — GSTR-3B
        hdr_fill_r = PatternFill("solid", fgColor="6e1a1a")   # dark red — GSTR-1
        hdr_fill_m = PatternFill("solid", fgColor="3c3c3c")   # dark grey — month sub-header
        center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left    = Alignment(horizontal="left",   vertical="center")
        thin    = Side(style="thin", color="CCCCCC")
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)

        filed_fill   = PatternFill("solid", fgColor="D4EDDA")
        pending_fill = PatternFill("solid", fgColor="FFF3CD")
        other_fill   = PatternFill("solid", fgColor="F0E6FF")

        filed_font   = Font(name="Arial", color="155724", size=9, bold=True)
        pending_font = Font(name="Arial", color="856404", size=9, bold=True)
        other_font   = Font(name="Arial", color="6b21a8", size=9, bold=True)
        data_font    = Font(name="Arial", size=9)

        # ── Row 1: main headers ──
        # Columns: 1=GSTIN, 2=Client Name,
        #          3-14 = GSTR-3B months, 15-26 = GSTR-1 months
        ws.merge_cells("A1:A2"); ws["A1"] = "GSTIN"
        ws.merge_cells("B1:B2"); ws["B1"] = "Client Name"
        ws.merge_cells("C1:N1"); ws["C1"] = "GSTR-3B"
        ws.merge_cells("O1:Z1"); ws["O1"] = "GSTR-1/IFF"

        for cell_ref, fill in [("A1",hdr_fill_b),("B1",hdr_fill_b),
                                 ("C1",hdr_fill_g),("O1",hdr_fill_r)]:
            ws[cell_ref].font      = hdr_font
            ws[cell_ref].fill      = fill
            ws[cell_ref].alignment = center
            ws[cell_ref].border    = border

        # Row 2 headers
        for i,m in enumerate(MONTHS_SHORT):
            # GSTR-3B months: cols C-N (3-14)
            c3b = ws.cell(row=2, column=3+i, value=m)
            c3b.font = hdr_font; c3b.fill = hdr_fill_m; c3b.alignment = center; c3b.border = border
            # GSTR-1 months: cols O-Z (15-26)
            cr1 = ws.cell(row=2, column=15+i, value=m)
            cr1.font = hdr_font; cr1.fill = hdr_fill_m; cr1.alignment = center; cr1.border = border

        for c_ref in ["A2","B2"]:
            ws[c_ref].fill   = hdr_fill_b
            ws[c_ref].border = border

        # ── Data rows ──
        for r_idx, row in enumerate(results_data, start=3):
            ws.cell(row=r_idx, column=1, value=row.get("GSTIN","")).alignment = left
            ws.cell(row=r_idx, column=2, value=row.get("Client Name","")).alignment = left
            for c in range(1,3):
                ws.cell(row=r_idx,column=c).font   = data_font
                ws.cell(row=r_idx,column=c).border = border

            for i, month in enumerate(MONTHS):
                # GSTR-3B
                val3b = row.get(f"3B_{MONTHS_SHORT[i]}", "Pending")
                cell3b = ws.cell(row=r_idx, column=3+i, value=val3b)
                cell3b.alignment = center; cell3b.border = border
                if val3b == "Filed":
                    cell3b.fill = filed_fill; cell3b.font = filed_font
                elif val3b == "Pending":
                    cell3b.fill = pending_fill; cell3b.font = pending_font
                else:
                    cell3b.fill = other_fill; cell3b.font = other_font

                # GSTR-1
                valr1 = row.get(f"R1_{MONTHS_SHORT[i]}", "Pending")
                cellr1 = ws.cell(row=r_idx, column=15+i, value=valr1)
                cellr1.alignment = center; cellr1.border = border
                if valr1 == "Filed":
                    cellr1.fill = filed_fill; cellr1.font = filed_font
                elif valr1 == "Pending":
                    cellr1.fill = pending_fill; cellr1.font = pending_font
                else:
                    cellr1.fill = other_fill; cellr1.font = other_font

        # ── Column widths ──
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 22
        for col_idx in range(3, 27):
            ws.column_dimensions[get_column_letter(col_idx)].width = 8

        ws.freeze_panes = "C3"   # Freeze gstin/name columns + headers
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

        wb.save(OUTPUT_FILE)
    except Exception as e:
        log(f"  ⚠️  save_report_now error: {e}")
        log(traceback.format_exc())
        # Emergency fallback — plain pandas save
        try:
            pd.DataFrame(results_data).to_excel(OUTPUT_FILE, index=False)
        except: pass

# ─────────────────────────────────────────────
# MAIN AUTOMATION
# ─────────────────────────────────────────────
def run_automation(fin_year, resume_from=0):
    try:
        _run_automation_inner(fin_year, resume_from)
    except Exception as e:
        log(f"❌ FATAL: {e}")
        log(traceback.format_exc())
        global is_done, session_active
        is_done = True; session_active = False

def _run_automation_inner(fin_year, resume_from=0):
    global results_data, is_done, skip_flag, current_client, progress, stop_flag, session_active

    if uploaded_df is None:
        log("❌ No Excel file uploaded."); is_done = True; session_active = False; return

    df = uploaded_df.copy().fillna("")

    # ── Column detection ──
    col_map = {}
    for col in df.columns:
        cl = col.strip().lower()
        if "gstin" in cl or "gst no" in cl or "gst number" in cl:
            col_map.setdefault("gstin", col)
        elif "client" in cl and "name" in cl:
            col_map.setdefault("client_name", col)

    missing = [k for k in ["gstin"] if k not in col_map]
    if missing:
        log(f"❌ Could not detect columns: {missing}. Found: {list(df.columns)}")
        is_done = True; session_active = False; return

    log(f"✅ Column map: {col_map}")

    progress["total"] = len(df)
    progress["done"]  = resume_from

    if resume_from > 0:
        log(f"↩  Resuming from client {resume_from+1} of {len(df)}...")
    else:
        log(f"📂 {len(df)} clients | Financial Year: {fin_year}")
        log(f"🚀 Opening Chrome...\n")

    driver = init_driver()

    df = df.reset_index(drop=True)  # ensure idx matches resume_from integer
    for idx, row in df.iterrows():
        if idx < resume_from: continue
        skip_flag = False

        if stop_flag:
            log("⏸  Stop flag — saving session...")
            save_session(idx, fin_year, progress["done"], len(df), results_data)
            save_report_now()
            break

        gstin       = str(row[col_map["gstin"]]).strip()
        client_name = str(row.get(col_map.get("client_name",""),"")).strip() or f"Client {idx+1}"

        current_client.update({"name":client_name,"gstin":gstin,"index":idx})
        log(f"\n{'─'*52}")
        log(f"[{idx+1}/{len(df)}] {client_name} | {gstin}")
        log(f"{'─'*52}")

        result_row = {
            "GSTIN"       : gstin,
            "Client Name" : client_name,
            "Financial Year": fin_year,
            "Processed At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Remarks"     : "",
        }
        # Initialise all month columns to Pending
        for ms in MONTHS_SHORT:
            result_row[f"3B_{ms}"] = "Pending"
            result_row[f"R1_{ms}"] = "Pending"

        statuses = fetch_filing_status(driver, gstin, fin_year)
        # Write month-wise results
        for i, month in enumerate(MONTHS):
            ms = MONTHS_SHORT[i]
            result_row[f"3B_{ms}"] = statuses["3B"].get(month, "Pending")
            result_row[f"R1_{ms}"] = statuses["R1"].get(month, "Pending")
        stopped = any(v=="Stopped" for v in statuses["3B"].values())
        skipped = any(v=="Skipped" for v in statuses["3B"].values())
        timed_out = any(v=="Captcha Timeout" for v in statuses["3B"].values())

        if stopped:
            # Session was stopped mid-client — don't record a result for this
            # GSTIN, and resume from this same client (retry it) next time.
            stop_flag = True
            log(f"  ⏸  Stop requested while processing {client_name} — will retry this client on resume.")
            save_session(idx, fin_year, progress["done"], len(df), results_data)
            save_report_now()
            break

        result_row["Remarks"] = (
            "Skipped" if skipped else
            "Captcha Timeout" if timed_out else
            "OK"
        )

        results_data.append(result_row)
        progress["done"] = idx + 1

        # ── Save after every client ──
        save_report_now()
        log(f"  💾 Report saved ({progress['done']} clients).")

        time.sleep(1)

    try:
        driver.quit()
    except:
        pass
    save_report_now()

    if not stop_flag:
        clear_session_file()
        log(f"\n{'='*52}")
        log(f"✅ ALL DONE! {len(results_data)} clients processed.")
        log(f"⬇️  Click the green Download button to save your report.")
        log(f"{'='*52}")
        is_done = True
        session_active = False
    else:
        log(f"⏸  Session paused. {progress['done']} of {len(df)} clients processed.")
        log(f"   Re-open and click '↩ Resume' to continue.")
        session_active = False

# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    port = 5050

    print(f"\n{'='*52}")
    print(f"  Filedge GST")
    print(f"  Opening at http://localhost:{port}")
    print(f"{'='*52}\n")

    threading.Timer(
        1.5,
        lambda: webbrowser.open(f"http://localhost:{port}")
    ).start()

    try:
        app.run(port=port, debug=False, use_reloader=False)
    except OSError:
        print("Port 5050 already in use.")